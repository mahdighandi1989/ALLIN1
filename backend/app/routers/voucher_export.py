"""v105 — Contra/security-cheque voucher: Excel template export.

The voucher page renders slips from a small "spec" (kind/title/date/GL/amount/
ref/description/name/extra lines) — the SAME values the print preview shows.
This endpoint turns that spec into a real .xlsx laid out like the printed
voucher (the form's original home was an Excel macro workbook), so the officer
can fill it later in Excel and print it looking the same:

  • merged lavender banner, bordered OUR REF box, signature rules,
    a thick outer frame per slip — all real cells/borders, editable;
  • A4 portrait print setup, fit-to-one-page, margins matching the web print
    (8mm top / ~11mm sides) so Excel's print mirrors the page's print;
  • the bank logo embedded top-right of each slip (sent by the client as the
    same bundled PNG the page shows).

Deterministic rendering with a real library (openpyxl) per the established
artifact philosophy — no AI, no side-channel storage; the response is the file.
"""
from __future__ import annotations

import base64
import io
import re
from typing import List, Optional

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import Response
from pydantic import BaseModel, Field

from app.routers.auth import get_current_active_user

router = APIRouter(tags=["voucher-export"], dependencies=[Depends(get_current_active_user)])

_MAX_LOGO_B64 = 400_000            # ~300 KB decoded — the bundled logo is far smaller


class SlipSpec(BaseModel):
    kind: str = Field(..., max_length=10)          # DEBIT | CREDIT
    title: str = Field("", max_length=60)          # banner text
    date: str = Field("", max_length=30)
    ac_no: str = Field("", max_length=60)
    amount: str = Field("", max_length=40)         # display text (already formatted)
    amount_boxed: bool = False                      # IRR: AMOUNT / QUANTITY box
    our_ref: str = Field("", max_length=120)
    description: str = Field("", max_length=200)
    ac_name: str = Field("", max_length=200)
    extra_lines: List[str] = Field(default_factory=list)


class VoucherExportIn(BaseModel):
    mode: str = Field("normal", max_length=12)
    slips: List[SlipSpec] = Field(..., min_length=1, max_length=4)
    logo_png: Optional[str] = Field(None, max_length=_MAX_LOGO_B64)   # data URL


def _decode_logo(data_url: str | None) -> bytes | None:
    if not data_url:
        return None
    m = re.match(r"^data:image/(png|jpeg);base64,([A-Za-z0-9+/=\s]+)$", data_url)
    if not m:
        return None
    try:
        raw = base64.b64decode(m.group(2), validate=False)
    except Exception:
        return None
    return raw if 0 < len(raw) <= 350_000 else None


def build_voucher_workbook(payload: VoucherExportIn) -> bytes:
    from openpyxl import Workbook
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.page import PageMargins

    wb = Workbook()
    ws = wb.active
    ws.title = "Voucher"

    # ~188 mm across 7 columns (Excel column width ≈ 2.03 mm/unit at Calibri 11;
    # Arial is close enough for a fit-to-width print)
    widths = [15, 13, 13, 13, 13, 12, 14]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    LAST = len(widths)                    # G
    LASTL = get_column_letter(LAST)

    thin = Side(style="thin", color="000000")
    medium = Side(style="medium", color="000000")
    thick = Side(style="thick", color="000000")
    lavender = PatternFill("solid", fgColor="CCCCFF")

    def box(r1: int, c1: int, r2: int, c2: int, side: Side) -> None:
        """Outline border around a rectangular range (per-edge, keeps insides)."""
        for c in range(c1, c2 + 1):
            top = ws.cell(row=r1, column=c)
            top.border = Border(top=side, left=top.border.left, right=top.border.right, bottom=top.border.bottom)
            bot = ws.cell(row=r2, column=c)
            bot.border = Border(bottom=side, left=bot.border.left, right=bot.border.right, top=bot.border.top)
        for r in range(r1, r2 + 1):
            lef = ws.cell(row=r, column=c1)
            lef.border = Border(left=side, top=lef.border.top, right=lef.border.right, bottom=lef.border.bottom)
            rig = ws.cell(row=r, column=c2)
            rig.border = Border(right=side, top=rig.border.top, left=rig.border.left, bottom=rig.border.bottom)

    logo = _decode_logo(payload.logo_png)
    row = 1
    for slip in payload.slips:
        top_row = row
        # -- header: kind (left) / INTERNAL VOUCHER (right; logo image above) --
        ws.row_dimensions[row].height = 40
        c = ws.cell(row=row, column=1, value=slip.kind.strip().upper()[:10])
        c.font = Font(name="Arial", size=26, bold=True)
        c.alignment = Alignment(horizontal="left", vertical="center")
        if logo:
            try:
                img = XLImage(io.BytesIO(logo))
                scale = 53.0 / max(1.0, float(img.height))     # ≈14 mm tall
                img.height = int(img.height * scale)
                img.width = int(img.width * scale)
                ws.add_image(img, f"E{row}")
            except Exception:
                pass
        row += 1
        ws.row_dimensions[row].height = 16
        c = ws.cell(row=row, column=LAST - 2, value="INTERNAL VOUCHER")
        ws.merge_cells(start_row=row, start_column=LAST - 2, end_row=row, end_column=LAST)
        c.font = Font(name="Arial", size=11, bold=True)
        c.alignment = Alignment(horizontal="right")
        row += 1

        # -- lavender banner --
        ws.row_dimensions[row].height = 24
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=LAST)
        c = ws.cell(row=row, column=1, value=slip.title.strip())
        c.font = Font(name="Arial", size=16, bold=True)
        c.alignment = Alignment(horizontal="center", vertical="center")
        for col in range(1, LAST + 1):
            ws.cell(row=row, column=col).fill = lavender
        box(row, 1, row, LAST, thin)
        row += 1

        # -- DATE (right) --
        ws.merge_cells(start_row=row, start_column=LAST - 2, end_row=row, end_column=LAST)
        c = ws.cell(row=row, column=LAST - 2, value=f"DATE :  {slip.date.strip()}")
        c.font = Font(name="Arial", size=11, bold=True)
        c.alignment = Alignment(horizontal="right")
        row += 2

        # -- A/c No (left) + amount (right) --
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        c = ws.cell(row=row, column=1, value=f"A/c No. :   {slip.ac_no.strip()}")
        c.font = Font(name="Arial", size=13, bold=True)
        c.alignment = Alignment(horizontal="left")
        if slip.amount_boxed:
            c = ws.cell(row=row, column=LAST - 2, value="AMOUNT / QUANTITY")
            c.font = Font(name="Arial", size=9, bold=True)
            c.alignment = Alignment(horizontal="right", vertical="center")
            ws.merge_cells(start_row=row, start_column=LAST - 2, end_row=row, end_column=LAST - 1)
            a = ws.cell(row=row, column=LAST, value=slip.amount.strip() or "—")
            a.font = Font(name="Arial", size=12, bold=True)
            a.alignment = Alignment(horizontal="center", vertical="center")
            box(row, LAST, row, LAST, medium)
        else:
            ws.merge_cells(start_row=row, start_column=LAST - 1, end_row=row, end_column=LAST)
            a = ws.cell(row=row, column=LAST - 1, value=slip.amount.strip() or "**********")
            a.font = Font(name="Arial", size=12, bold=True)
            a.alignment = Alignment(horizontal="right")
        row += 2

        # -- OUR REF box --
        ref_top = row
        c = ws.cell(row=row, column=1, value="OUR REF :")
        c.font = Font(name="Arial", size=10, bold=True)
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=LAST)
        c = ws.cell(row=row, column=2, value=slip.our_ref.strip())
        c.font = Font(name="Arial", size=11, bold=True)
        row += 1
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=LAST)
        c = ws.cell(row=row, column=2, value=slip.description.strip())
        c.font = Font(name="Arial", size=10)
        row += 1
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=LAST)
        c = ws.cell(row=row, column=2, value=slip.ac_name.strip())
        c.font = Font(name="Arial", size=10, bold=True)
        for col in range(2, LAST + 1):   # the name line's top rule (like the form)
            cell = ws.cell(row=row, column=col)
            cell.border = Border(top=thin, left=cell.border.left, right=cell.border.right, bottom=cell.border.bottom)
        row += 1
        for ln in slip.extra_lines[:6]:
            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=LAST)
            c = ws.cell(row=row, column=2, value=str(ln)[:200])
            c.font = Font(name="Arial", size=10, bold=True)
            row += 1
        ref_bot = row - 1
        # label column separator + box outline
        for r in range(ref_top, ref_bot + 1):
            cell = ws.cell(row=r, column=1)
            cell.border = Border(right=medium, top=cell.border.top, left=cell.border.left, bottom=cell.border.bottom)
        box(ref_top, 1, ref_bot, LAST, medium)
        row += 2

        # -- signature footer --
        ws.row_dimensions[row].height = 34
        row += 1
        c = ws.cell(row=row, column=1, value="Prepared By.")
        c.font = Font(name="Arial", size=10, bold=True)
        c = ws.cell(row=row, column=LAST - 2, value="Authorized Signatures")
        c.font = Font(name="Arial", size=10, bold=True)
        row += 1
        for col in (1, 2):
            cell = ws.cell(row=row, column=col)
            cell.border = Border(top=thin, left=cell.border.left, right=cell.border.right, bottom=cell.border.bottom)
        for col in (LAST - 2, LAST - 1, LAST):
            cell = ws.cell(row=row, column=col)
            cell.border = Border(top=thin, left=cell.border.left, right=cell.border.right, bottom=cell.border.bottom)
        row += 1

        # -- thick outer frame around the whole slip --
        box(top_row, 1, row - 1, LAST, thick)
        row += 2                          # gap before the next slip

    # -- A4 print setup mirroring the web print (one page, safe margins) --
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins = PageMargins(left=0.43, right=0.43, top=0.31, bottom=0.35, header=0.1, footer=0.1)
    ws.print_area = f"A1:{LASTL}{row}"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@router.post("/export-excel")
async def export_voucher_excel(payload: VoucherExportIn):
    try:
        data = build_voucher_workbook(payload)
    except Exception as exc:
        raise HTTPException(status_code=422, detail=f"ساختِ فایلِ Excel ناموفق بود: {exc}")
    fname = f"Voucher-{(payload.mode or 'normal').strip() or 'normal'}.xlsx"
    return Response(
        content=data,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )
