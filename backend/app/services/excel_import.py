"""Parse an uploaded spreadsheet into customer/facility rows.

Reuses the robust reader idea from ``app.services.data_pipeline`` but works on
in-memory bytes (an uploaded file). Returns parsed dict rows + a header list so
the import router can map and validate them. Pure parsing — no DB access.

Format support (detected by magic bytes, not the filename, so a mislabelled
upload is still handled correctly):

* ``.xlsx`` / ``.xlsm`` — Office Open XML (a zip archive) read via ``openpyxl``.
* ``.xls`` — the legacy OLE2 binary format read via ``xlrd`` (when installed).

Every failure raises :class:`ExcelParseError` carrying a ``kind`` so callers can
turn it into a precise, user-facing message instead of a generic 500.
"""
from __future__ import annotations

import io
import logging
from typing import Dict, Iterable, List, Sequence, Tuple

logger = logging.getLogger("app.imports")

# Magic byte signatures used to dispatch to the right reader.
_ZIP_MAGIC = b"PK\x03\x04"  # .xlsx / .xlsm (Office Open XML == zip archive)
_OLE2_MAGIC = b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"  # legacy .xls (OLE2 compound)

# Kinds of parse failure (mirrors data_pipeline.PipelineError.kind).
ERROR_KINDS = ("empty", "corrupt", "no_sheets", "unsupported", "too_many_columns", "error")


class ExcelParseError(Exception):
    """Raised when the uploaded file cannot be read as a spreadsheet.

    ``kind`` is one of :data:`ERROR_KINDS` and lets the caller produce a precise
    message (e.g. distinguish an *empty* file from a *corrupt* one).
    """

    def __init__(self, message: str, *, kind: str = "error") -> None:
        super().__init__(message)
        self.kind = kind


def _detect_format(content: bytes) -> str | None:
    """Return 'xlsx', 'xls', or None based on the leading magic bytes."""
    if content[:4] == _ZIP_MAGIC:
        return "xlsx"
    if content[:8] == _OLE2_MAGIC:
        return "xls"
    return None


def _rows_to_records(
    rows_iter: Iterable[Sequence], *, max_rows: int, max_cols: int = 256
) -> Tuple[List[str], List[Dict]]:
    """Shared logic: turn an iterator of raw row tuples into (headers, records).

    The first non-empty row is the header. Header cells are lowercased and
    stripped so callers can match columns case-insensitively. Fully-empty rows
    are skipped; at most ``max_rows`` data rows are returned.
    """
    header = None
    for row in rows_iter:
        if row and any(c is not None and str(c).strip() != "" for c in row):
            header = row
            break
    if header is None:
        raise ExcelParseError("the spreadsheet has no header row or data", kind="empty")

    if len(header) > max_cols:
        raise ExcelParseError(
            f"too many columns ({len(header)} > {max_cols})", kind="too_many_columns"
        )

    headers = [
        (str(h).strip().lower() if h is not None else f"col_{i}")
        for i, h in enumerate(header)
    ]

    # Duplicate header names would silently collapse columns (the right-most
    # wins in the dict below) — a copy-pasted second "amount" column would
    # import the wrong values into every row. Fail fast instead.
    seen: Dict = {}
    dupes = []
    for h in headers:
        seen[h] = seen.get(h, 0) + 1
        if seen[h] == 2:
            dupes.append(h)
    if dupes:
        raise ExcelParseError(
            f"duplicate column name(s): {', '.join(sorted(dupes))} — "
            "rename or remove the repeated column(s) and re-upload",
            kind="duplicate_columns",
        )

    records: List[Dict] = []
    for row in rows_iter:
        if row is None or all(c is None for c in row):
            continue
        if len(records) >= max_rows:
            # Never truncate silently: the operator would see "5000 rows
            # imported, 0 errors" and believe the whole file is in.
            raise ExcelParseError(
                f"the sheet has more than {max_rows} data rows; split the "
                f"file and import it in parts (limit per upload: {max_rows})",
                kind="too_many_rows",
            )
        rec = {key: (row[i] if i < len(row) else None) for i, key in enumerate(headers)}
        records.append(rec)
    return headers, records


def _read_xlsx(content: bytes, *, max_rows: int) -> Tuple[List[str], List[Dict]]:
    """Read the first sheet of an .xlsx/.xlsm byte stream via openpyxl."""
    import openpyxl  # lazy import

    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:  # corrupt / not actually a zip workbook
        raise ExcelParseError(f"could not read workbook: {exc}", kind="corrupt")
    try:
        if not wb.sheetnames:
            raise ExcelParseError("workbook has no sheets", kind="no_sheets")
        ws = wb[wb.sheetnames[0]]
        return _rows_to_records(ws.iter_rows(values_only=True), max_rows=max_rows)
    finally:
        wb.close()


def _read_xls(content: bytes, *, max_rows: int) -> Tuple[List[str], List[Dict]]:
    """Read the first sheet of a legacy binary .xls byte stream via xlrd."""
    try:
        import xlrd  # lazy import; only needed for legacy .xls
    except ImportError:
        raise ExcelParseError(
            "legacy .xls files are not supported on this server (xlrd missing); "
            "please re-save the file as .xlsx",
            kind="unsupported",
        )
    try:
        book = xlrd.open_workbook(file_contents=content)
    except Exception as exc:
        raise ExcelParseError(f"could not read .xls workbook: {exc}", kind="corrupt")
    if book.nsheets == 0:
        raise ExcelParseError("workbook has no sheets", kind="no_sheets")
    sheet = book.sheet_by_index(0)
    rows = (sheet.row_values(r) for r in range(sheet.nrows))
    return _rows_to_records(rows, max_rows=max_rows)


def parse_workbook(content: bytes, *, max_rows: int = 5000) -> Tuple[List[str], List[Dict]]:
    """Parse an uploaded spreadsheet (``.xlsx``/``.xlsm``/``.xls``) byte stream.

    Returns ``(headers, rows)`` where ``headers`` are normalised (lowercased,
    stripped) column names and ``rows`` is a list of dicts keyed by those
    headers. Raises :class:`ExcelParseError` (with a ``kind``) on any failure.
    """
    if not content:
        raise ExcelParseError("the uploaded file is empty", kind="empty")

    fmt = _detect_format(content)
    if fmt == "xlsx":
        return _read_xlsx(content, max_rows=max_rows)
    if fmt == "xls":
        return _read_xls(content, max_rows=max_rows)
    raise ExcelParseError(
        "unrecognised file format — expected an .xlsx, .xlsm or .xls workbook",
        kind="corrupt",
    )


def validate_required_columns(headers: Sequence[str], required: Sequence[str]) -> List[str]:
    """Return the required columns that are missing from ``headers`` (case-insensitive).

    Lets a router fail fast with a precise "missing column" message instead of
    reporting the same per-row error on every single row.
    """
    present = {h.strip().lower() for h in headers}
    return [c for c in required if c.strip().lower() not in present]


def cell_str(value) -> str:
    """Normalise a cell value to a trimmed string ('' for None)."""
    if value is None:
        return ""
    # xlrd surfaces integers stored in .xls as floats (e.g. 5.0); trim a
    # trailing ".0" so account numbers / codes round-trip cleanly.
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()
