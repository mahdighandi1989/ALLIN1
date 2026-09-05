"""Drive «policy inbox» (v117) — pure logic.

The owner drops a batch of ISSUED insurance-policy scans plus one mapping
Excel into a dedicated Drive folder. Each policy page carries printed
identifiers (کد رایانهٔ بیمه‌نامه، شمارهٔ بیمه‌نامه، کد یکتای بیمه مرکزی،
کد ملی مالک/راهن) but NO bank account number; the Excel maps those
identifiers to the bank's branch + account. The app reads each policy's
identifiers with one small multimodal call, finds the UNIQUE matching Excel
row, and renames the Drive file to ``{branch}-{account}-{original name}`` —
which makes the import path's filename-based account attribution (v85)
deterministic.

This module holds only the deterministic, testable pieces: workbook parsing
with flexible Persian headers, digit normalization, unique matching (never
guess on ambiguity), and name construction. Drive/API/model calls live in
``routers/policy_inbox.py``.
"""
from __future__ import annotations

import re
from typing import Any, Dict, List, Optional, Tuple

FOLDER_NAME = "بیمه‌نامه‌ها - ورودی"

# Persian + Arabic-Indic digits → ASCII, so «۴۹۸۸۵۲۹» matches "4988529".
_DIGITS = str.maketrans("۰۱۲۳۴۵۶۷۸۹٠١٢٣٤٥٦٧٨٩", "01234567890123456789")

# Column-header synonyms (substring match on the normalized header text).
_HEADERS: Dict[str, Tuple[str, ...]] = {
    "account": ("شماره حساب", "شماره  حساب", "حساب", "account"),
    "branch": ("شماره شعبه", "کد شعبه", "شعبه", "branch"),
    "computer_code": ("کد رایانه", "كد رايانه", "computer"),
    "policy_no": ("شماره بیمه", "شماره بيمه", "بیمه نامه", "بیمه‌نامه", "policy"),
    "unique_code": ("کد یکتا", "كد يكتا", "یکتا", "unique"),
    "national_id": ("کد ملی", "كد ملي", "شناسه ملی", "national"),
    "name": ("نام مشتری", "نام", "customer", "مشتری"),
}
# Identifier precedence for matching a policy to a row (most specific first).
MATCH_KEYS = ("computer_code", "unique_code", "policy_no", "national_id")
_MIN_KEY_DIGITS = 4  # anything shorter is too weak to match on


def norm_digits(v: Any) -> str:
    """Digits only, Persian/Arabic translated to ASCII — the matching key."""
    s = str(v if v is not None else "")
    if s.endswith(".0"):  # openpyxl floats like 4988529.0
        s = s[:-2]
    return re.sub(r"\D", "", s.translate(_DIGITS))


def _norm_header(v: Any) -> str:
    s = str(v if v is not None else "").strip().lower()
    # unify the Persian/Arabic yeh & kaf variants + zwnj/spaces
    return (s.replace("ي", "ی").replace("ك", "ک").replace("‌", " ")
            .replace("ـ", "").strip())


def _header_key(cell: Any) -> Optional[str]:
    h = _norm_header(cell)
    if not h:
        return None
    for key, needles in _HEADERS.items():
        if any(n in h for n in needles):
            return key
    return None


def parse_mapping_workbook(data: bytes) -> Dict[str, Any]:
    """Parse the mapping Excel → {ok, rows, columns, warnings}.

    Finds the header row anywhere in the first 12 rows (owner tables often have
    a title row above the headers). Each data row becomes
    ``{account, branch, computer_code, policy_no, unique_code, national_id, name}``
    (missing cells empty). Rows without an account number are dropped with a
    warning — they cannot drive a rename."""
    import io

    from openpyxl import load_workbook

    warnings: List[str] = []
    try:
        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001
        return {"ok": False, "error": f"فایل اکسل قابلِ خواندن نبود: {exc}", "rows": [], "warnings": []}
    ws = wb.active
    grid = [[c for c in row] for row in ws.iter_rows(values_only=True)]
    wb.close()

    header_at, col_map = -1, {}
    for i, row in enumerate(grid[:12]):
        m = {}
        for j, cell in enumerate(row or []):
            k = _header_key(cell)
            if k and k not in m:
                m[k] = j
        # a usable header row names at least the account column + one match key
        if "account" in m and any(k in m for k in MATCH_KEYS):
            header_at, col_map = i, m
            break
    if header_at < 0:
        return {"ok": False, "rows": [], "warnings": [],
                "error": ("سطرِ عنوان پیدا نشد — جدول باید ستونِ «شماره حساب» و دست‌کم یکی از "
                          "«کد رایانه / کد یکتا / شماره بیمه‌نامه / کد ملی» را داشته باشد.")}

    rows: List[Dict[str, str]] = []
    dropped = 0
    for row in grid[header_at + 1:]:
        if not row or all(c in (None, "") for c in row):
            continue
        item: Dict[str, str] = {}
        for key, j in col_map.items():
            v = row[j] if j < len(row) else None
            s = str(v).strip() if v is not None else ""
            if s.endswith(".0"):
                s = s[:-2]
            item[key] = s
        if not norm_digits(item.get("account")):
            dropped += 1
            continue
        rows.append(item)
    if dropped:
        warnings.append(f"{dropped} ردیفِ بدونِ شماره حساب نادیده گرفته شد")
    if not rows:
        return {"ok": False, "rows": [], "warnings": warnings,
                "error": "هیچ ردیفِ دارای شماره حساب در اکسل نبود."}
    return {"ok": True, "rows": rows, "columns": sorted(col_map), "warnings": warnings}


def match_row(ids: Dict[str, str], rows: List[Dict[str, str]]) -> Tuple[Optional[Dict[str, str]], str]:
    """Find the UNIQUE Excel row for a policy's printed identifiers.

    Tries each identifier in :data:`MATCH_KEYS` order; the first identifier
    that yields exactly ONE row wins. Ambiguity or no-hit falls through to the
    next key; if nothing decides, (None, reason) — never a guess."""
    tried: List[str] = []
    for key in MATCH_KEYS:
        v = norm_digits(ids.get(key))
        if len(v) < _MIN_KEY_DIGITS:
            continue
        hits = [r for r in rows if norm_digits(r.get(key)) == v]
        if len(hits) == 1:
            return hits[0], key
        tried.append(f"{key}={v}({len(hits)})")
    if not tried:
        return None, "روی بیمه‌نامه هیچ شناسهٔ قابلِ‌تطبیقی خوانده نشد"
    return None, "تطبیقِ یکتا پیدا نشد: " + "، ".join(tried)


def already_named(name: str) -> bool:
    """True when the file already carries the ``{branch}-{account}-`` prefix
    (rename is idempotent — never stack prefixes)."""
    return bool(re.match(r"^\d{1,6}-\d{4,15}-", (name or "").strip()))


def build_new_name(branch: str, account: str, original: str) -> str:
    """Owner rule: new file name = branch number + account number + current name."""
    b = norm_digits(branch)
    a = norm_digits(account)
    orig = (original or "").strip() or "بیمه‌نامه.pdf"
    prefix = f"{b}-{a}-" if b else f"{a}-"
    return (prefix + orig)[:200]
