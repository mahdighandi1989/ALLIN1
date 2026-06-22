"""AI document-ingestion core.

Sends an uploaded file (PDF / image / …) to a vision/document model with a strong
extraction prompt, parses the strict JSON it returns, and persists each customer's
data — deduped — exactly like the rest of the app:
  • scalar facts → promoted profile columns + data_json (keyed, overwrite-in-place)
  • a credit-review row when the doc is a committee approval (deduped per date)
  • guarantors via upsert (which feeds the cross-account relationship graph)
  • a page→document map kept under the profile + on the attachment link

The file itself is stored in Google Drive by the caller (so the main DB stays
light); only metadata/links are kept in the DB.
"""
from __future__ import annotations

import json
import re
from datetime import date

from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.customer import Customer, AccountType
from app.models.crm import CustomerProfile
from app.models.guarantor import Guarantor
from app.models.profile_entities import MortgagedProperty
from app.services.customer_link import ensure_customer

# ---------------------------------------------------------------------------
# Schema-driven field registry. The list of facts the model is asked to extract
# (and that ``persist_customer`` writes to real columns) is DERIVED from the live
# ``CustomerProfile`` schema — so when a new column is added to the model later,
# it is automatically asked-for and persisted with NO change here. Only the few
# non-extractable columns below are excluded.
# ---------------------------------------------------------------------------
_PROFILE_SKIP_COLS = {
    # identity / handled at the customer level
    "account_no", "customer_name", "account_type", "branch",
    # housekeeping / computed
    "customer_status", "profile_completeness", "updated_by",
    "last_updated", "data_json", "created_at",
    # per-document file paths (populated by the upload feature, not from content)
    "trade_license_doc", "passport_doc", "emirates_id_doc", "visa_doc", "tenancy_doc",
    # officer-only free-text notes — never auto-filled by the model
    "trade_license_remarks", "passport_remarks", "emirates_id_remarks",
}
# Friendlier key the model is more likely to emit  ->  real column it maps to.
_FIELD_ALIASES = {"nationality": "passport_nationality"}
_COL_TO_FRIENDLY = {v: k for k, v in _FIELD_ALIASES.items()}


def extractable_profile_fields() -> list[str]:
    """The field keys the model should fill, derived live from the CustomerProfile
    schema (with a friendly alias where one exists). Adding a column to the model
    automatically adds it here — no prompt or persist edits needed."""
    out: list[str] = []
    for c in CustomerProfile.__table__.columns:
        if c.name in _PROFILE_SKIP_COLS:
            continue
        out.append(_COL_TO_FRIENDLY.get(c.name, c.name))
    return out


def _build_fields_block() -> str:
    """Pretty 3-per-line JSON skeleton of the extractable fields for the prompt."""
    parts = [f'"{k}": ""' for k in extractable_profile_fields()]
    return ",\n".join("        " + ", ".join(parts[i:i + 3]) for i in range(0, len(parts), 3))


def build_extraction_prompt() -> str:
    """Assemble the extraction prompt with the field list taken from the schema."""
    return (
        """You are a meticulous UAE banking credit-file analyst. Read the ATTACHED document(s) end to end (every page) and extract the data for the bank's customer database.

A single file MAY contain data for MORE THAN ONE customer/account (e.g. a borrower plus guarantors, or several approvals). Detect every distinct account and attribute each fact to the CORRECT account — never mix customers.

Return STRICT JSON ONLY (no markdown, no commentary), exactly this shape:
{
  "customers": [
    {
      "account_no": "<6-digit core account number>",
      "account_display": "<full as printed, e.g. 2624-115524-011>",
      "name": "<customer/company name>",
      "account_type": "retail" | "corporate",
      "branch": "<branch name and/or code>",
      "fields": {
"""
        + _build_fields_block()
        + """
      },
      "guarantors": [ {"name": "", "account": "", "branch": ""} ],
      "partners": [ {"name": "", "nationality": "", "share": "", "remarks": ""} ],
      "facilities": [ {"facility_type": "overdraft | loan | cheque_discounting | trust_receipt | lc_sight | lc_usance | lc | lg | log | other",
                       "amount": "", "currency": "AED", "interest_rate": "", "expiry_date": "", "notes": ""} ],
      "properties": [ {"prop_type": "", "address": "", "city": "", "country": "",
                       "valuation": "", "valuation_currency": "AED", "mortgage_amount": "", "mortgage_currency": "AED",
                       "plate_no": "", "mortgage_deed_no": "", "mortgage_date": "", "insurance_expiry": ""} ],
      "security": [ {"type": "Underlien Deposits | Cheques | Collaterals | <as printed>", "for_facility": "",
                     "aed": "", "usd": "", "irr": "", "other": ""} ],
      "review": {
        "date_of_review": "", "credit_application_no": "", "purpose": "",
        "proposed_rating": "", "rating_notes": "", "cru_recommendation": ""
      }
    }
  ],
  "documents": [
    {"pages": "1-2", "type": "Trade License", "customer_account": "<6-digit>", "summary": "<one line>"}
  ]
}

Rules:
- Extract EVERY field listed under "fields" that the document actually contains — including the small ones officers often miss: every ID's issue date AND expiry date, visa number/issue/expiry/type, tenancy number/issue/expiry/address, whether the Emirates ID is "golden" (Yes/No), etc.
- "partners" = the company's shareholders/partners (name, nationality, share %). "guarantors" = people/companies guaranteeing the facility. They are DIFFERENT — never confuse them.
- "facilities" = EVERY credit facility / limit (overdraft, loan, cheque discounting, trust receipt, LC sight/usance, LG, letter of guarantee, …) with its amount/limit, interest rate or margin, and expiry. Map each to the closest facility_type above; use "other" only if none fits.
- "security" = the collateral/security matrix: underlien deposits, security cheques, collaterals, etc., with the amount in each currency column (AED/USD/IRR/other) and which facility it secures ("for_facility").
- "grade" = the customer's history grade (VERY GOOD / GOOD / AVERAGE / POOR). "call_report" and "previous_files" (No. of Previous Files) come from the summary header. "undertaking_from" = who gives undertaking forms (e.g. "Guarantor/s", "Partner/s"). Fill these whenever the file shows them.
- Only include fields you actually find; omit unknowns (do NOT invent values).
- Always capture the customer's full legal NAME exactly as printed.
- Capture each ID document's NUMBER together with its ISSUE and EXPIRY dates.
- "nationality" ONLY if it is explicitly printed — never guess or assume it.
- Dates as printed (DD/MM/YYYY is fine); a RENEWED document's later date should be reported so the record updates.
- "account_no" is the 6-digit core (the middle group of 2624-115524-011 is 115524).
- Numbers as plain digits (no thousands separators) where possible.
- "documents" must list, for EVERY page or page-range, which document it is and which account it belongs to.
- Output ONLY the JSON object.""")


# Built once from the current schema (the model is fully defined at import).
EXTRACTION_PROMPT = build_extraction_prompt()


def parse_model_json(text: str) -> dict:
    """Best-effort parse of the model's JSON reply (tolerates code fences/prose)."""
    if not text:
        return {}
    t = text.strip()
    t = re.sub(r"^```(?:json)?", "", t).strip()
    if t.endswith("```"):
        t = t[:-3].strip()
    i, j = t.find("{"), t.rfind("}")
    if i >= 0 and j > i:
        t = t[i:j + 1]
    try:
        out = json.loads(t)
        return out if isinstance(out, dict) else {}
    except Exception:
        return {}


def _newer_or_empty(new: str, old: str) -> bool:
    """True if ``new`` should replace ``old`` for a date field: old empty, or new
    is a later (renewed) date. Unparseable new → keep old (don't clobber)."""
    if not (new or "").strip():
        return False
    if not (old or "").strip():
        return True
    try:
        from dateutil import parser as _dp
        nd = _dp.parse(new, dayfirst=True, fuzzy=True).date()
    except Exception:
        return False
    try:
        from dateutil import parser as _dp
        od = _dp.parse(old, dayfirst=True, fuzzy=True).date()
    except Exception:
        return True
    return nd >= od


_KYC_DATE_COLS = {
    "trade_license_issue", "trade_license_expiry", "passport_issue", "passport_expiry",
    "emirates_id_issue", "emirates_id_expiry", "visa_expiry", "tenancy_expiry",
}


def _acc_of(cust: dict) -> str:
    acc = str(cust.get("account_no") or "").strip()
    if re.fullmatch(r"\d{6}", acc):
        return acc
    for src in (acc, str(cust.get("account_display") or "")):
        m = re.search(r"\b(\d{6})\b", src)
        if m:
            return m.group(1)
    return acc


_SEC_COLS = ("type", "for_facility", "aed", "usd", "irr", "other")


def _merge_security(data: dict, security: list) -> int:
    """Upsert the security/collateral matrix into ``data['security_details']``,
    deduped by row type (fill-empty per currency column). Returns rows added."""
    rows = [s for s in (security or []) if isinstance(s, dict)
            and any(str(s.get(k) or "").strip() for k in _SEC_COLS)]
    if not rows:
        return 0
    out = data.get("security_details")
    if not isinstance(out, list):
        out = []
    added = 0
    for s in rows:
        stype = str(s.get("type") or "").strip()
        match = next((e for e in out if isinstance(e, dict)
                      and stype and str(e.get("type") or "").strip().lower() == stype.lower()), None)
        if match is None:
            out.append({k: str(s.get(k) or "").strip() for k in _SEC_COLS})
            added += 1
        else:
            for k in _SEC_COLS:
                v = str(s.get(k) or "").strip()
                if v and not str(match.get(k) or "").strip():
                    match[k] = v
    data["security_details"] = out
    return added


def _apply_extracted_fields(cp, fields: dict) -> None:
    """Write extracted scalars onto the profile, SCHEMA-DRIVEN: any key matching a
    CustomerProfile column (directly or via alias) is promoted to that column.

    Write policy (smart, no blind overwrite):
      • KYC document dates (issue/expiry) UPDATE on renewal — a later date wins, so
        a renewed passport/EID/visa/tenancy refreshes the record.
      • everything else is fill-empty — a possibly-wrong extraction never clobbers
        a value an officer has already curated.
    """
    cols = {c.name: c for c in CustomerProfile.__table__.columns}
    for key, v in (fields or {}).items():
        if v in (None, ""):
            continue
        col = _FIELD_ALIASES.get(key, key)
        if col in _PROFILE_SKIP_COLS or col not in cols:
            continue  # not a writable column → stays in data_json only
        cur = getattr(cp, col, "")
        cur = cur.strip() if isinstance(cur, str) else (cur or "")
        take = _newer_or_empty(str(v), str(cur)) if col in _KYC_DATE_COLS else (not cur)
        if take:
            ml = getattr(getattr(cols[col], "type", None), "length", None)
            setattr(cp, col, str(v)[:ml] if ml else str(v))


async def persist_customer(db: AsyncSession, cust: dict, username: str, source: str = "import_ai") -> dict:
    """Apply one extracted customer dict to the DB (deduped). Returns a summary."""
    from app.routers.crm import _upsert_credit_review  # shared helper

    acc = _acc_of(cust)
    if not acc:
        return {"ok": False, "reason": "no_account"}
    name = (cust.get("name") or "").strip() or None
    customer = await ensure_customer(db, acc, name)
    at = str(cust.get("account_type") or "").lower()
    if customer is not None and at in ("retail", "corporate", "sme"):
        customer.account_type = AccountType(at)
    if customer is not None and name and (not customer.name or customer.name.strip() in ("", acc)):
        customer.name = name[:200]
    branch = (cust.get("branch") or "").strip()
    if customer is not None and branch and not (customer.branch or "").strip():
        customer.branch = branch[:100]

    fields = {k: v for k, v in (cust.get("fields") or {}).items() if v not in (None, "")}

    cp = (await db.execute(select(CustomerProfile).where(CustomerProfile.account_no == acc))).scalar_one_or_none()
    if cp is None:
        cp = CustomerProfile(account_no=acc, customer_name=name)
        db.add(cp)
    # Schema-driven: promote EVERY recognised field to its real column (identity,
    # KYC docs incl. issue dates / visa type / tenancy address / golden EID, and
    # the credit scalars). Adding a new column to the model later is picked up here
    # automatically — no edit needed.
    _apply_extracted_fields(cp, fields)
    try:
        data = json.loads(cp.data_json) if cp.data_json else {}
        if not isinstance(data, dict):
            data = {}
    except Exception:
        data = {}
    for k, v in fields.items():
        data[str(k)] = v  # keyed → no duplicates on re-import
    # Security/collateral matrix → data_json["security_details"] (the credit-file
    # form reads this). Upsert by type so a re-import refreshes rows in place and
    # never duplicates them; amounts are fill-empty.
    sec_added = _merge_security(data, cust.get("security") or [])
    cp.data_json = json.dumps(data, ensure_ascii=False)
    cp.last_updated = date.today().isoformat()
    cp.updated_by = username

    review = {k: v for k, v in (cust.get("review") or {}).items() if v not in (None, "")}
    if review:
        review.setdefault("customer_name", name)
        if at:
            review.setdefault("account_type", at)
        await _upsert_credit_review(db, acc, review, source, username)

    g_added = g_updated = 0
    for g in (cust.get("guarantors") or []):
        gname = (g.get("name") or "").strip()
        gacc = (g.get("account") or "").strip()
        if not gname:
            continue
        if gacc:
            await ensure_customer(db, gacc, gname)
        row = None
        if gacc:
            row = (await db.execute(select(Guarantor).where(
                Guarantor.account_no == acc, Guarantor.guarantor_account == gacc,
                Guarantor.is_deleted == False))).scalar_one_or_none()  # noqa: E712
        if row is None:
            row = (await db.execute(select(Guarantor).where(
                Guarantor.account_no == acc, Guarantor.guarantor_name == gname,
                Guarantor.is_deleted == False))).scalar_one_or_none()  # noqa: E712
        if row is None:
            import uuid
            from datetime import datetime
            row = Guarantor(id=f"G-{acc}-{datetime.now().strftime('%Y%m%d%H%M%S')}-{uuid.uuid4().hex[:2]}",
                            account_no=acc, date_added=date.today().isoformat(), created_by=username)
            db.add(row)
            g_added += 1
        else:
            g_updated += 1
        row.guarantor_name = gname[:200]
        if gacc:
            row.guarantor_account = gacc[:50]
        if g.get("branch"):
            row.branch = str(g["branch"])[:20]
        if customer is not None and customer.name and not row.customer_name:
            row.customer_name = customer.name

    # Partners / shareholders (شرکا) — upsert into the Partner table the corporate
    # credit-file form reads from. Deduped by name within the account; fill-empty
    # so a curated nationality/share is never clobbered. (Distinct from guarantors.)
    from app.models.profile_entities import Partner
    pt_added = pt_updated = 0
    existing_partners = (await db.execute(select(Partner).where(
        Partner.account_no == acc, Partner.is_deleted == False))).scalars().all()  # noqa: E712
    for pt in (cust.get("partners") or []):
        if not isinstance(pt, dict):
            continue
        pname = (pt.get("name") or "").strip()
        if not pname:
            continue
        prow = next((r for r in existing_partners if (r.name or "").strip().lower() == pname.lower()), None)
        if prow is None:
            import uuid
            from datetime import datetime
            prow = Partner(id=f"PT-{acc}-{datetime.now().strftime('%Y%m%d%H%M%S')}-{uuid.uuid4().hex[:2]}",
                           account_no=acc, date_added=date.today().isoformat(), created_by=username)
            db.add(prow)
            existing_partners.append(prow)
            pt_added += 1
        else:
            pt_updated += 1
        prow.name = pname[:200]
        _set_prop(prow, "nationality", pt.get("nationality"))
        _set_prop(prow, "share", pt.get("share"))
        _set_prop(prow, "remarks", pt.get("remarks"))
        if customer is not None and customer.name and not prow.customer_name:
            prow.customer_name = customer.name

    # Properties (املاک) — upsert with smart matching so a similar property isn't
    # duplicated: match on plate no / mortgage deed no, else type+address.
    p_added = p_updated = 0
    for p in (cust.get("properties") or []):
        if not isinstance(p, dict):
            continue
        ptype = (p.get("prop_type") or "").strip()
        addr = (p.get("address") or "").strip()
        plate = (p.get("plate_no") or "").strip()
        deed = (p.get("mortgage_deed_no") or "").strip()
        if not (ptype or addr or plate or deed):
            continue
        prow = await _match_property(db, acc, plate, deed, ptype, addr)
        if prow is None:
            import uuid
            from datetime import datetime
            prow = MortgagedProperty(id=f"PROP-{acc}-{datetime.now().strftime('%Y%m%d%H%M%S')}-{uuid.uuid4().hex[:2]}",
                                     account_no=acc, date_added=date.today().isoformat(), created_by=username)
            db.add(prow)
            p_added += 1
        else:
            p_updated += 1
        # fill-empty so curated values are never clobbered
        _set_prop(prow, "prop_type", ptype)
        _set_prop(prow, "address", addr)
        _set_prop(prow, "city", p.get("city"))
        _set_prop(prow, "country", p.get("country"))
        _set_prop(prow, "plate_no", plate)
        _set_prop(prow, "mortgage_deed_no", deed)
        _set_prop(prow, "mortgage_date", p.get("mortgage_date"))
        _set_prop(prow, "insurance_expiry", p.get("insurance_expiry"))
        _set_prop(prow, "valuation_currency", p.get("valuation_currency"))
        _set_prop(prow, "mortgage_currency", p.get("mortgage_currency"))
        if prow.valuation in (None, "") and _num_bounded(p.get("valuation"), 1e16) is not None:
            prow.valuation = _num_bounded(p.get("valuation"), 1e16)  # Numeric(18,2)
        if prow.mortgage_amount in (None, "") and _num_bounded(p.get("mortgage_amount"), 1e16) is not None:
            prow.mortgage_amount = _num_bounded(p.get("mortgage_amount"), 1e16)
        if customer is not None and customer.name and not prow.customer_name:
            prow.customer_name = customer.name

    # Facilities — create/upsert the real Facility records the credit-file form
    # reads (so Facility Details fills in). Matched by (customer, facility_type);
    # amounts/rate/expiry are fill-empty (a curated amount is never clobbered) and
    # a NEW record is only created when an amount is present (no phantom limits).
    f_added = f_updated = 0
    if customer is not None:
        from app.models.facility import Facility, FacilityType, FacilityStatus
        valid_ft = {t.value for t in FacilityType}
        existing_facs = (await db.execute(select(Facility).where(
            Facility.customer_id == customer.id, Facility.is_deleted == False))).scalars().all()  # noqa: E712
        for fc in (cust.get("facilities") or []):
            if not isinstance(fc, dict):
                continue
            ft_raw = (fc.get("facility_type") or "").strip().lower().replace(" ", "_")
            ft = ft_raw if ft_raw in valid_ft else ("other" if ft_raw else "")
            amt = _num_bounded(fc.get("amount"), 1e13)        # Numeric(15,2)
            rate = _num_bounded(fc.get("interest_rate"), 1e3)  # Numeric(5,2) → < 1000
            frow = None
            if ft:
                frow = next((r for r in existing_facs
                             if str(getattr(r.facility_type, "value", r.facility_type) or "") == ft), None)
            if frow is None:
                if amt is None:
                    continue  # no amount → nothing to anchor a new facility on
                frow = Facility(customer_id=customer.id,
                                facility_type=FacilityType(ft) if ft else FacilityType.OTHER,
                                amount=amt, currency=(fc.get("currency") or "AED")[:3].upper(),
                                status=FacilityStatus.ACTIVE)
                db.add(frow)
                existing_facs.append(frow)
                f_added += 1
            else:
                if not frow.amount and amt is not None:
                    frow.amount = amt
                f_updated += 1
            if rate is not None and not frow.interest_rate:
                frow.interest_rate = rate
            if fc.get("notes") and not (frow.notes or ""):
                frow.notes = str(fc["notes"])
            exp = _parse_date(fc.get("expiry_date"))
            if exp and frow.expiry_date is None:
                frow.expiry_date = exp

    return {"ok": True, "account_no": acc, "name": name or acc,
            "customer_id": (customer.id if customer is not None else None),
            "facility_hint": (cust.get("fields", {}).get("proposed_facility")
                              or (cust.get("review", {}) or {}).get("proposed_facility") or ""),
            "fields_saved": sorted(fields.keys()),
            "guarantors_added": g_added, "guarantors_updated": g_updated,
            "partners_added": pt_added, "partners_updated": pt_updated,
            "facilities_added": f_added, "facilities_updated": f_updated,
            "properties_added": p_added, "properties_updated": p_updated,
            "security_added": sec_added}


def _num(v):
    try:
        s = re.sub(r"[^\d.\-]", "", str(v or ""))
        return float(s) if s not in ("", "-", ".") else None
    except Exception:
        return None


def _num_bounded(v, max_abs: float):
    """Parse a number but DROP it if its magnitude won't fit the target DB column.

    A mis-extracted value (e.g. an interest_rate of 3190 into a Numeric(5,2)
    column) would otherwise raise a numeric-overflow on flush and roll back the
    WHOLE import. Returning None here just skips that one value."""
    n = _num(v)
    if n is None or abs(n) >= max_abs:
        return None
    return n


def _parse_date(v):
    """Parse a printed date (DD/MM/YYYY etc.) to a date, or None if unparseable."""
    s = str(v or "").strip()
    if not s:
        return None
    try:
        from dateutil import parser as _dp
        return _dp.parse(s, dayfirst=True, fuzzy=True).date()
    except Exception:
        return None


def _set_prop(row, col: str, v) -> None:
    """Fill-empty a string column on a property row (don't clobber)."""
    v = (str(v).strip() if v not in (None, "") else "")
    if v and not (getattr(row, col, "") or "").strip():
        column = row.__table__.columns.get(col)
        ml = getattr(getattr(column, "type", None), "length", None)
        setattr(row, col, v[:ml] if ml else v)


async def _match_property(db: AsyncSession, acc: str, plate: str, deed: str, ptype: str, addr: str):
    """Find an existing property for ``acc`` matching plate/deed, else type+address."""
    from app.models.profile_entities import MortgagedProperty as MP
    rows = (await db.execute(select(MP).where(MP.account_no == acc, MP.is_deleted == False))).scalars().all()  # noqa: E712
    for r in rows:
        if plate and (r.plate_no or "").strip().lower() == plate.lower():
            return r
        if deed and (r.mortgage_deed_no or "").strip().lower() == deed.lower():
            return r
    if ptype and addr:
        for r in rows:
            if (r.prop_type or "").strip().lower() == ptype.lower() and (r.address or "").strip().lower() == addr.lower():
                return r
    return None


def record_documents_on_profile(data: dict, documents: list, drive_link: str, drive_id: str, filename: str, sha: str = "") -> dict:
    """Append the page→document map + Drive link to a profile data_json dict
    (deduped by content hash, falling back to drive id). Mutates and returns it."""
    if not isinstance(data, dict):
        data = {}
    docs = data.get("imported_documents")
    if not isinstance(docs, list):
        docs = []
    # Drop any prior entry for the SAME file (by content hash or Drive id) so a
    # re-import refreshes in place instead of duplicating.
    def _same(d):
        if not isinstance(d, dict):
            return False
        if sha and d.get("sha") == sha:
            return True
        return bool(drive_id) and d.get("drive_id") == drive_id
    docs = [d for d in docs if not _same(d)]
    docs.append({
        "filename": filename, "drive_id": drive_id, "drive_link": drive_link, "sha": sha,
        "pages": [d for d in (documents or []) if isinstance(d, dict)],
        "imported_at": date.today().isoformat(),
    })
    data["imported_documents"] = docs
    return data


# ---------------------------------------------------------------------------
# Spreadsheet / Office tables → text, so a big table spanning many accounts can
# be fed to the model (chunked) and every row routed to its customer.
# ---------------------------------------------------------------------------
def workbook_to_text(data: bytes, filename: str) -> str:
    """Flatten an Excel/CSV file to CSV-like text across ALL sheets."""
    name = (filename or "").lower()
    if name.endswith(".csv"):
        for enc in ("utf-8-sig", "utf-8", "latin-1"):
            try:
                return data.decode(enc)
            except Exception:
                continue
        return data.decode("utf-8", "replace")
    out: list[str] = []
    if name.endswith((".xlsx", ".xlsm")):
        import io
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        try:
            for sh in wb.sheetnames:
                ws = wb[sh]
                out.append(f"## Sheet: {sh}")
                for row in ws.iter_rows(values_only=True):
                    cells = ["" if v is None else str(v) for v in row]
                    if any(c.strip() for c in cells):
                        out.append(",".join(cells))
        finally:
            wb.close()
    elif name.endswith(".xls"):
        import xlrd
        wb = xlrd.open_workbook(file_contents=data)
        for sh in wb.sheets():
            out.append(f"## Sheet: {sh.name}")
            for r in range(sh.nrows):
                cells = [str(sh.cell_value(r, c)) for c in range(sh.ncols)]
                if any(c.strip() for c in cells):
                    out.append(",".join(cells))
    return "\n".join(out)


def chunk_text(text: str, max_chars: int = 100000) -> list[str]:
    """Split table text into chunks under ``max_chars``, repeating the sheet/header
    hint at the top of each chunk so every chunk stays self-describing."""
    lines = (text or "").split("\n")
    header = "\n".join(lines[:2])
    chunks: list[str] = []
    cur: list[str] = []
    size = 0
    for ln in lines:
        if size + len(ln) + 1 > max_chars and cur:
            chunks.append("\n".join(cur))
            cur = [header] if header else []
            size = len(header)
        cur.append(ln)
        size += len(ln) + 1
    if cur:
        chunks.append("\n".join(cur))
    return chunks or [text]


def _lc(d: dict, k: str) -> str:
    return (d.get(k) or "").strip().lower()


def _g_match(a: dict, b: dict) -> bool:
    if _lc(a, "account") and _lc(a, "account") == _lc(b, "account"):
        return True
    return bool(_lc(b, "name")) and _lc(a, "name") == _lc(b, "name")


def _name_match(a: dict, b: dict) -> bool:
    return bool(_lc(b, "name")) and _lc(a, "name") == _lc(b, "name")


def _ft_match(a: dict, b: dict) -> bool:
    return bool(_lc(b, "facility_type")) and _lc(a, "facility_type") == _lc(b, "facility_type")


def _sec_match(a: dict, b: dict) -> bool:
    return bool(_lc(b, "type")) and _lc(a, "type") == _lc(b, "type")


def _prop_match(a: dict, b: dict) -> bool:
    for key in ("plate_no", "mortgage_deed_no"):
        if _lc(a, key) and _lc(a, key) == _lc(b, key):
            return True
    # same type and (one side has no address yet, or addresses agree) → same property
    if _lc(a, "prop_type") and _lc(a, "prop_type") == _lc(b, "prop_type"):
        aa, ab = _lc(a, "address"), _lc(b, "address")
        if not aa or not ab or aa == ab:
            return True
    return False


_LIST_MATCHERS = {"guarantors": _g_match, "partners": _name_match,
                  "facilities": _ft_match, "properties": _prop_match, "security": _sec_match}


def _merge_list(into_list: list, more_list: list, matcher) -> None:
    """Merge ``more_list`` into ``into_list``: a matching item is field-merged
    (fill-empty) so a record split across chunks is reassembled; otherwise it is
    appended."""
    for item in (more_list or []):
        if not isinstance(item, dict):
            continue
        match = next((ex for ex in into_list if isinstance(ex, dict) and matcher(ex, item)), None)
        if match is None:
            into_list.append(item)
        else:
            for k, vv in item.items():
                if vv not in (None, "") and not match.get(k):
                    match[k] = vv


def merge_customer(into: dict, more: dict) -> None:
    """Merge a customer dict parsed from a later chunk into an existing one — so a
    big PDF split into page-chunks reassembles every list (partners, facilities,
    properties, guarantors), even when a record's details span two chunks."""
    for k, v in more.items():
        if k in ("fields", "review") and isinstance(v, dict):
            base = into.setdefault(k, {})
            for kk, vv in v.items():
                if vv not in (None, "") and not base.get(kk):
                    base[kk] = vv
        elif k in _LIST_MATCHERS and isinstance(v, list):
            into.setdefault(k, [])
            _merge_list(into[k], v, _LIST_MATCHERS[k])
        elif v not in (None, "") and not into.get(k):
            into[k] = v


# ---------------------------------------------------------------------------
# Oversized PDFs → split into page-chunks (each under the provider's inline
# limit) so a big file is extracted section-by-section and merged.
# ---------------------------------------------------------------------------
def split_pdf(data: bytes, max_bytes: int = 18 * 1024 * 1024, max_pages: int = 12):
    """Return ([(start_page_1based, pdf_bytes), ...], total_pages). Each chunk is
    <= max_bytes (or a single page if one page alone exceeds it)."""
    import io
    from pypdf import PdfReader, PdfWriter

    reader = PdfReader(io.BytesIO(data))
    n = len(reader.pages)
    chunks: list[tuple[int, bytes]] = []
    i = 0
    while i < n:
        end = min(i + max_pages, n)
        while end > i:
            w = PdfWriter()
            for p in range(i, end):
                w.add_page(reader.pages[p])
            buf = io.BytesIO()
            w.write(buf)
            b = buf.getvalue()
            if len(b) <= max_bytes or (end - i) == 1:
                chunks.append((i + 1, b))
                i = end
                break
            end -= 1
    return chunks, n


def offset_pages(pages: str, offset: int) -> str:
    """Shift the page numbers in a "1-2" / "3" string by ``offset`` (chunk → global)."""
    if not offset:
        return str(pages or "")
    return re.sub(r"\d+", lambda m: str(int(m.group()) + offset), str(pages or ""))
