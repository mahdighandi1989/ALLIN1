"""Data-file scan/verify pipeline for the imported Excel datasets.

This is the failure point that emits the critical ``scan_failed`` /
``verify_failed`` / ``task_failed`` notifications. ``run_scan_task`` is designed
to be safe to call from a scheduler/Celery worker: it never raises, instead
reporting failures through :func:`app.services.notifications.notify_event`.
"""
from __future__ import annotations

import logging
import os
from dataclasses import dataclass, field
from typing import List, Optional

from app.services.notifications import notify_event

logger = logging.getLogger(__name__)

# Default location of the source spreadsheets (repo-relative ``data-import/``).
DEFAULT_DATA_DIR = os.path.join(
    os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(__file__)))),
    "data-import",
)

_SPREADSHEET_EXTS = (".xlsx", ".xlsm", ".xls")


@dataclass
class ScanResult:
    scanned: List[str] = field(default_factory=list)
    failed: List[str] = field(default_factory=list)
    ok: bool = True
    error: Optional[str] = None


def _load_workbook(path: str):
    """Lazily import openpyxl so the module imports even if it is absent."""
    import openpyxl  # imported lazily on purpose

    return openpyxl.load_workbook(path, read_only=True, data_only=True)


def scan_data_files(data_dir: str = DEFAULT_DATA_DIR) -> ScanResult:
    """Open every spreadsheet under ``data_dir`` and confirm it is readable.

    On any failure a high-priority, non-silent ``scan_failed`` notification is
    emitted (in Persian) and the result is marked not-ok.
    """
    result = ScanResult()

    if not os.path.isdir(data_dir):
        result.ok = False
        result.error = f"directory not found: {data_dir}"
        # notify_event("scan_failed", ...) with silent=False, priority="high"
        notify_event("scan_failed", f"پوشهٔ دادهٔ ورودی یافت نشد: {data_dir}", silent=False, priority="high")
        return result

    files = [
        f for f in sorted(os.listdir(data_dir)) if f.lower().endswith(_SPREADSHEET_EXTS)
    ]

    for name in files:
        path = os.path.join(data_dir, name)
        try:
            wb = _load_workbook(path)
            _ = wb.sheetnames
            wb.close()
            result.scanned.append(name)
        except Exception as exc:  # corrupt/unreadable spreadsheet
            logger.error("Data scan failed for %s: %s", name, exc)
            result.failed.append(name)
            result.ok = False
            notify_event(
                "scan_failed",
                f"خواندن فایل دادهٔ «{name}» با خطا مواجه شد: {exc}",
                silent=False,
                priority="high",
            )

    return result


def verify_scan_result(result: ScanResult) -> bool:
    """Verify that the scan produced usable data; emit ``verify_failed`` if not."""
    if not result.ok or not result.scanned:
        notify_event(
            "verify_failed",
            f"راستی‌آزمایی اسکن ناموفق بود؛ فایل‌های ناموفق: {result.failed or 'هیچ فایلی اسکن نشد'}",
            silent=False,
            priority="high",
        )
        return False
    return True


def run_scan_task(data_dir: str = DEFAULT_DATA_DIR) -> ScanResult:
    """Run the full scan+verify as a background task.

    Wraps everything so an unexpected error still surfaces a ``task_failed``
    notification instead of crashing the worker.
    """
    try:
        result = scan_data_files(data_dir)
        verify_scan_result(result)
        return result
    except Exception as exc:  # pragma: no cover - defensive catch-all
        logger.error("Scan task crashed: %s", exc, exc_info=True)
        notify_event(
            "task_failed",
            f"اجرای وظیفهٔ اسکن دادهٔ سیستم به‌طور غیرمنتظره متوقف شد: {exc}",
            silent=False,
            priority="high",
        )
        return ScanResult(ok=False, error=str(exc))
