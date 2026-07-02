"""Document/data exporters: CSV (stdlib) and PDF (reportlab, with HTML fallback).

Kept dependency-tolerant: if reportlab is unavailable the PDF helpers return a
printable HTML document instead of crashing, so the endpoints always respond.
"""
from __future__ import annotations

import csv
import html
import io
from typing import Any, Dict, List, Sequence, Tuple

try:  # reportlab is optional; degrade to HTML if missing.
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        SimpleDocTemplate,
        Paragraph,
        Spacer,
        Table,
        TableStyle,
    )

    _HAS_REPORTLAB = True
except Exception:  # pragma: no cover - exercised only when reportlab absent
    _HAS_REPORTLAB = False


# --- CSV --------------------------------------------------------------------
def _csv_safe(v: Any) -> Any:
    """Neutralize spreadsheet formula injection in user-entered text.

    The BOM below makes Excel the expected consumer, and Excel executes cells
    starting with = + - @ (e.g. a customer name imported as
    '=HYPERLINK(...)'). Prefix a single quote so Excel treats it as text.
    Numbers and other non-strings pass through untouched.
    """
    if isinstance(v, str) and v[:1] in ("=", "+", "-", "@"):
        return "'" + v
    return v


def rows_to_csv(headers: Sequence[str], rows: Sequence[Sequence[Any]]) -> bytes:
    """Render a header + rows into UTF-8 CSV bytes (Excel-friendly BOM)."""
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(list(headers))
    for row in rows:
        writer.writerow(["" if v is None else _csv_safe(v) for v in row])
    # Prepend a BOM so Excel opens UTF-8 (e.g. Arabic names) correctly.
    return ("﻿" + buf.getvalue()).encode("utf-8")


# --- PDF --------------------------------------------------------------------
def _html_fallback(title: str, sections: List[Tuple[str, Sequence[str], Sequence[Sequence[Any]]]],
                   meta: Dict[str, Any] | None = None) -> bytes:
    parts = [
        "<!doctype html><html><head><meta charset='utf-8'>",
        f"<title>{html.escape(title)}</title>",
        "<style>body{font-family:Arial,sans-serif;margin:32px;color:#111}"
        "h1{font-size:20px}h2{font-size:15px;margin-top:24px}"
        "table{border-collapse:collapse;width:100%;margin-top:8px;font-size:12px}"
        "th,td{border:1px solid #ccc;padding:6px;text-align:left}"
        "th{background:#f3f4f6}.meta{color:#555;font-size:12px}</style></head><body>",
        f"<h1>{html.escape(title)}</h1>",
    ]
    if meta:
        parts.append("<p class='meta'>")
        parts.append(" &nbsp;·&nbsp; ".join(f"{html.escape(str(k))}: {html.escape(str(v))}" for k, v in meta.items()))
        parts.append("</p>")
    for heading, headers, rows in sections:
        parts.append(f"<h2>{html.escape(heading)}</h2>")
        parts.append("<table><thead><tr>")
        parts += [f"<th>{html.escape(str(h))}</th>" for h in headers]
        parts.append("</tr></thead><tbody>")
        for row in rows:
            parts.append("<tr>")
            parts += [f"<td>{html.escape('' if v is None else str(v))}</td>" for v in row]
            parts.append("</tr>")
        parts.append("</tbody></table>")
    parts.append("</body></html>")
    return "".join(parts).encode("utf-8")


def build_pdf(
    title: str,
    sections: List[Tuple[str, Sequence[str], Sequence[Sequence[Any]]]],
    meta: Dict[str, Any] | None = None,
) -> Tuple[bytes, str]:
    """Build a document from titled table sections.

    Returns ``(content_bytes, media_type)``. When reportlab is available this is a
    real PDF; otherwise it is printable HTML (so the caller still gets a usable
    document and the right ``Content-Disposition`` extension).
    """
    if not _HAS_REPORTLAB:
        return _html_fallback(title, sections, meta), "text/html"

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        leftMargin=18 * mm, rightMargin=18 * mm, topMargin=18 * mm, bottomMargin=18 * mm,
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("t", parent=styles["Title"], fontSize=18, spaceAfter=6)
    meta_style = ParagraphStyle("m", parent=styles["Normal"], fontSize=9, textColor=colors.grey)
    h2 = ParagraphStyle("h2", parent=styles["Heading2"], fontSize=12, spaceBefore=12)

    elements: List[Any] = [Paragraph(html.escape(title), title_style)]
    if meta:
        meta_line = " &nbsp; · &nbsp; ".join(
            f"<b>{html.escape(str(k))}</b>: {html.escape(str(v))}" for k, v in meta.items()
        )
        elements.append(Paragraph(meta_line, meta_style))
    elements.append(Spacer(1, 6))

    for heading, headers, rows in sections:
        elements.append(Paragraph(html.escape(heading), h2))
        data = [list(headers)] + [["" if v is None else str(v) for v in r] for r in rows]
        table = Table(data, repeatRows=1, hAlign="LEFT")
        table.setStyle(
            TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1f4e8c")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#cccccc")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f5f7fb")]),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("LEFTPADDING", (0, 0), (-1, -1), 5),
                ("RIGHTPADDING", (0, 0), (-1, -1), 5),
                ("TOPPADDING", (0, 0), (-1, -1), 3),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ])
        )
        elements.append(table)

    doc.build(elements)
    return buf.getvalue(), "application/pdf"


# --- XLSX -------------------------------------------------------------------
XLSX_MEDIA_TYPE = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)


def _safe_sheet_title(title: str, used: set) -> str:
    """Excel sheet titles: <=31 chars, no []:*?/\\ , and unique."""
    cleaned = "".join("_" if c in "[]:*?/\\" else c for c in (title or "Sheet"))[:31] or "Sheet"
    base = cleaned
    i = 2
    while cleaned.lower() in used:
        suffix = f"_{i}"
        cleaned = base[: 31 - len(suffix)] + suffix
        i += 1
    used.add(cleaned.lower())
    return cleaned


def build_xlsx(
    sections: List[Tuple[str, Sequence[str], Sequence[Sequence[Any]]]],
) -> bytes:
    """Build a real .xlsx workbook: one styled sheet per (title, headers, rows).

    Numeric-looking cells are written as numbers so Excel sums/sorts them; the
    header row is bold on a banded fill with frozen panes and auto-ish widths.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    wb.remove(wb.active)  # drop the default empty sheet
    used_titles: set = set()
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E8C")

    if not sections:
        sections = [("Sheet1", ["(empty)"], [])]

    for title, headers, rows in sections:
        ws = wb.create_sheet(_safe_sheet_title(title, used_titles))
        ws.append(list(headers))
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(vertical="center")

        widths = [len(str(h)) for h in headers]
        for row in rows:
            out_row = []
            for i, v in enumerate(row):
                num = _coerce_number(v)
                out_row.append(num if num is not None else ("" if v is None else v))
                if i < len(widths):
                    widths[i] = max(widths[i], len(str(v if v is not None else "")))
            ws.append(out_row)

        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = min(60, max(10, w + 2))
        ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _coerce_number(v: Any):
    """Return v as int/float if it cleanly represents a number, else None.

    Strips thousands separators and a leading currency code (e.g. 'AED 1,234.50').
    """
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return v
    if isinstance(v, str):
        s = v.strip().replace(",", "")
        # Drop a leading non-numeric currency prefix like "AED ".
        parts = s.split()
        if len(parts) == 2 and not parts[0].replace(".", "").isdigit():
            s = parts[1]
        # Identifiers must survive the round trip to Excel as TEXT:
        # a leading zero (account "012345", phone "0501234567") would be
        # destroyed by int coercion, and very long digit runs render as
        # float/scientific notation. Only coerce values that read back
        # identically as numbers.
        if s.startswith("0") and s != "0" and not s.startswith("0."):
            return None
        try:
            if s and (s.lstrip("-").isdigit()):
                if len(s.lstrip("-")) > 15:  # beyond exact float/display range
                    return None
                return int(s)
            f = float(s)
            return f
        except (ValueError, AttributeError):
            return None
    return None
