"""Robust Excel data pipeline.

This pipeline turns the source spreadsheets under ``data-import/`` into structured
records for downstream consumers (CSV export / database loading) while:

* **Error handling** — corrupt, empty, or invalid-format files are detected and
  handled (logged + a ``scan_failed`` notification) instead of crashing the
  pipeline (see ``process_file``: every read is wrapped in try/except and raises
  a clear :class:`PipelineError`).
* **Schema validation** — each sheet is validated against an expected
  :class:`SheetSchema` (required columns). Note an .xlsx/.xlsm file is really a
  binary archive file (a zip of XML parts); .xls is the legacy binary format —
  the reader is selected per extension.
* **Defined output** — the component output is an explicit list of row records
  (dicts); the original Excel file is preserved (opened read-only, never
  mutated) and the extracted data can be written downstream as CSV / loaded into
  the database.
* **Format support** — .xlsx, .xlsm and .xls are all supported.
"""
from __future__ import annotations

import csv
import logging
import os
from dataclasses import dataclass, field
from typing import Dict, List, Optional

from app.services.notifications import notify_event

logger = logging.getLogger(__name__)

# Supported spreadsheet formats. .xlsx/.xlsm are read via openpyxl; the legacy
# binary .xls format is read via xlrd when available.
SUPPORTED_EXTENSIONS = {".xlsx", ".xlsm", ".xls"}

DEFAULT_DATA_DIR = os.path.join(
    os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(__file__)))),
    "data-import",
)


class PipelineError(Exception):
    """Raised for an unrecoverable problem with a single source file."""

    def __init__(self, message: str, *, kind: str = "error"):
        super().__init__(message)
        self.kind = kind  # one of: corrupt, empty, invalid_format, schema, error


@dataclass
class SheetSchema:
    """Expected schema for a sheet: its (case-insensitive) required columns."""

    required_columns: List[str] = field(default_factory=list)


@dataclass
class FileResult:
    path: str
    ok: bool = True
    rows: List[Dict] = field(default_factory=list)
    error: Optional[str] = None
    error_kind: Optional[str] = None


@dataclass
class PipelineReport:
    results: List[FileResult] = field(default_factory=list)

    @property
    def ok(self) -> bool:
        return all(r.ok for r in self.results)

    @property
    def total_rows(self) -> int:
        return sum(len(r.rows) for r in self.results)


def _ext(path: str) -> str:
    return os.path.splitext(path)[1].lower()


def _read_xlsx_rows(path: str) -> List[Dict]:
    """Read the first sheet of an .xlsx/.xlsm (binary archive) into dict rows."""
    import openpyxl  # lazy import

    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:  # unreadable -> corrupt
        raise PipelineError(f"corrupt workbook: {exc}", kind="corrupt")

    try:
        if not wb.sheetnames:
            raise PipelineError("workbook has no sheets (empty)", kind="empty")
        ws = wb[wb.sheetnames[0]]
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header = next(rows_iter)
        except StopIteration:
            raise PipelineError("sheet is empty (no header row)", kind="empty")
        headers = [str(h) if h is not None else f"col_{i}" for i, h in enumerate(header)]
        records: List[Dict] = []
        for row in rows_iter:
            if row is None or all(c is None for c in row):
                continue
            records.append({headers[i]: row[i] if i < len(row) else None for i in range(len(headers))})
        return records
    finally:
        wb.close()


def _read_xls_rows(path: str) -> List[Dict]:
    """Read the first sheet of a legacy binary .xls file via xlrd."""
    try:
        import xlrd  # lazy import; only needed for .xls
    except ImportError:
        raise PipelineError(
            "the legacy .xls format requires the optional 'xlrd' package",
            kind="invalid_format",
        )
    try:
        book = xlrd.open_workbook(path)
    except Exception as exc:
        raise PipelineError(f"corrupt .xls workbook: {exc}", kind="corrupt")
    if book.nsheets == 0:
        raise PipelineError("workbook has no sheets (empty)", kind="empty")
    sheet = book.sheet_by_index(0)
    if sheet.nrows == 0:
        raise PipelineError("sheet is empty (no header row)", kind="empty")
    headers = [str(sheet.cell_value(0, c)) or f"col_{c}" for c in range(sheet.ncols)]
    records: List[Dict] = []
    for r in range(1, sheet.nrows):
        records.append({headers[c]: sheet.cell_value(r, c) for c in range(sheet.ncols)})
    return records


def load_rows(path: str) -> List[Dict]:
    """Load row records from a spreadsheet, dispatching by file extension."""
    if not os.path.exists(path):
        raise PipelineError(f"file not found: {path}", kind="error")
    ext = _ext(path)
    if ext not in SUPPORTED_EXTENSIONS:
        raise PipelineError(f"invalid format: {ext}", kind="invalid_format")
    if ext in (".xlsx", ".xlsm"):
        return _read_xlsx_rows(path)
    return _read_xls_rows(path)


def validate_schema(rows: List[Dict], schema: Optional[SheetSchema]) -> None:
    """Validate that ``rows`` contain the schema's required columns."""
    if schema is None or not schema.required_columns:
        return
    if not rows:
        raise PipelineError("schema validation failed: no data rows", kind="schema")
    present = {str(k).strip().lower() for k in rows[0].keys()}
    missing = [c for c in schema.required_columns if c.strip().lower() not in present]
    if missing:
        raise PipelineError(
            f"schema validation failed: missing columns {missing}", kind="schema"
        )


def process_file(path: str, schema: Optional[SheetSchema] = None) -> FileResult:
    """Process a single source file with full error handling.

    Never raises: corrupt/empty/invalid-format/schema problems are caught and
    reported on the :class:`FileResult` (and a ``scan_failed`` notification is
    emitted). The original Excel file is never modified.
    """
    result = FileResult(path=path)
    try:
        rows = load_rows(path)
        validate_schema(rows, schema)
        result.rows = rows
        result.ok = True
    except PipelineError as exc:
        result.ok = False
        result.error = str(exc)
        result.error_kind = exc.kind
        logger.error("Pipeline error for %s (%s): %s", path, exc.kind, exc)
        notify_event(
            "scan_failed",
            f"پردازش فایل «{os.path.basename(path)}» ناموفق بود ({exc.kind}): {exc}",
            silent=False,
            priority="high",
        )
    except Exception as exc:  # defensive: unexpected error
        result.ok = False
        result.error = str(exc)
        result.error_kind = "error"
        logger.error("Unexpected pipeline error for %s: %s", path, exc, exc_info=True)
        notify_event(
            "scan_failed",
            f"خطای غیرمنتظره در پردازش فایل «{os.path.basename(path)}»: {exc}",
            silent=False,
            priority="high",
        )
    return result


def export_to_csv(rows: List[Dict], output_path: str) -> str:
    """Write extracted records downstream as CSV (original Excel is untouched)."""
    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    fieldnames: List[str] = []
    for row in rows:
        for key in row.keys():
            if key not in fieldnames:
                fieldnames.append(key)
    with open(output_path, "w", newline="", encoding="utf-8") as fh:
        writer = csv.DictWriter(fh, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow(row)
    return output_path


def run_pipeline(
    data_dir: str = DEFAULT_DATA_DIR,
    schema: Optional[SheetSchema] = None,
    output_dir: Optional[str] = None,
) -> PipelineReport:
    """Process every supported spreadsheet under ``data_dir``.

    Output: a :class:`PipelineReport` of per-file extracted records. Original
    Excel files are preserved (read-only). When ``output_dir`` is given, each
    file's records are also exported downstream as CSV.
    """
    report = PipelineReport()
    if not os.path.isdir(data_dir):
        logger.error("Data dir not found: %s", data_dir)
        notify_event("scan_failed", f"پوشهٔ داده یافت نشد: {data_dir}", silent=False, priority="high")
        return report

    for name in sorted(os.listdir(data_dir)):
        if _ext(name) not in SUPPORTED_EXTENSIONS:
            continue
        path = os.path.join(data_dir, name)
        result = process_file(path, schema)
        report.results.append(result)
        if result.ok and output_dir:
            try:
                export_to_csv(result.rows, os.path.join(output_dir, name + ".csv"))
            except Exception as exc:  # pragma: no cover - export is best-effort
                logger.warning("CSV export failed for %s: %s", name, exc)

    return report
