"""Tests for real .xlsx (and list CSV) exports."""
import io

import pytest
from httpx import AsyncClient

from app.models.customer import Customer, AccountType, CustomerStatus

XLSX_CT = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


class TestXlsxExports:
    async def test_requires_auth(self, client: AsyncClient):
        assert (await client.get("/api/customers/export.xlsx")).status_code == 401
        assert (await client.get("/api/reports/portfolio/export.xlsx")).status_code == 401

    async def test_customers_xlsx_is_real_workbook(
        self, client: AsyncClient, auth_headers: dict, db_session
    ):
        db_session.add_all([
            Customer(account_no="XL-1", name="Excel Co One", account_type=AccountType.CORPORATE, status=CustomerStatus.ACTIVE),
            Customer(account_no="XL-2", name="Excel Co Two", account_type=AccountType.RETAIL, status=CustomerStatus.ACTIVE),
        ])
        await db_session.commit()

        r = await client.get("/api/customers/export.xlsx", headers=auth_headers)
        assert r.status_code == 200
        assert r.headers["content-type"].startswith(XLSX_CT)
        assert r.content[:2] == b"PK"  # zip/xlsx magic

        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(r.content))
        ws = wb["Customers"]
        assert [c.value for c in ws[1]][:3] == ["id", "account_no", "name"]
        # 2 seeded rows + header
        assert ws.max_row >= 3

    async def test_customers_xlsx_respects_filter(
        self, client: AsyncClient, auth_headers: dict, db_session
    ):
        db_session.add_all([
            Customer(account_no="F-CORP", name="FilterCorp", account_type=AccountType.CORPORATE, status=CustomerStatus.ACTIVE),
            Customer(account_no="F-RET", name="FilterRetail", account_type=AccountType.RETAIL, status=CustomerStatus.ACTIVE),
        ])
        await db_session.commit()

        r = await client.get("/api/customers/export.xlsx?account_type=corporate", headers=auth_headers)
        import openpyxl
        ws = openpyxl.load_workbook(io.BytesIO(r.content))["Customers"]
        names = [ws.cell(row=i, column=3).value for i in range(2, ws.max_row + 1)]
        assert "FilterCorp" in names
        assert "FilterRetail" not in names

    async def test_facilities_xlsx(self, client: AsyncClient, auth_headers: dict, test_facility):
        r = await client.get("/api/facilities/export.xlsx", headers=auth_headers)
        assert r.status_code == 200 and r.content[:2] == b"PK"

    async def test_facilities_export_honours_sort(
        self, client: AsyncClient, auth_headers: dict, db_session, test_customer
    ):
        """export.xlsx must honour sort_by/sort_order so the file matches the
        on-screen list order (previously it was hard-coded to created_at desc)."""
        from app.models.facility import Facility, FacilityType, FacilityStatus

        db_session.add_all([
            Facility(customer_id=test_customer.id, name="Bravo", facility_type=FacilityType.LOAN,
                     status=FacilityStatus.ACTIVE, amount=300, outstanding=0, currency="AED"),
            Facility(customer_id=test_customer.id, name="Alpha", facility_type=FacilityType.LOAN,
                     status=FacilityStatus.ACTIVE, amount=100, outstanding=0, currency="AED"),
            Facility(customer_id=test_customer.id, name="Charlie", facility_type=FacilityType.LOAN,
                     status=FacilityStatus.ACTIVE, amount=200, outstanding=0, currency="AED"),
        ])
        await db_session.commit()

        import openpyxl

        def names(content):
            ws = openpyxl.load_workbook(io.BytesIO(content))["Facilities"]
            header = [c.value for c in ws[1]]
            col = header.index("name") + 1
            return [ws.cell(row=i, column=col).value for i in range(2, ws.max_row + 1)]

        r = await client.get(
            "/api/facilities/export.xlsx?sort_by=name&sort_order=asc", headers=auth_headers
        )
        assert r.status_code == 200
        assert names(r.content) == ["Alpha", "Bravo", "Charlie"]

        r = await client.get(
            "/api/facilities/export.xlsx?sort_by=name&sort_order=desc", headers=auth_headers
        )
        assert r.status_code == 200
        assert names(r.content) == ["Charlie", "Bravo", "Alpha"]

    async def test_portfolio_xlsx_multisheet(self, client: AsyncClient, auth_headers: dict):
        r = await client.get("/api/reports/portfolio/export.xlsx", headers=auth_headers)
        assert r.status_code == 200 and r.content[:2] == b"PK"
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(r.content))
        assert "Summary" in wb.sheetnames
        assert "Top Exposures" in wb.sheetnames

    async def test_customers_csv_export(self, client: AsyncClient, auth_headers: dict, db_session):
        db_session.add(Customer(account_no="CSV-1", name="Csv Co", account_type=AccountType.SME, status=CustomerStatus.ACTIVE))
        await db_session.commit()
        r = await client.get("/api/customers/export.csv", headers=auth_headers)
        assert r.status_code == 200
        assert "Csv Co" in r.content.decode("utf-8")

    async def test_xlsx_numbers_are_numeric(self, client: AsyncClient, auth_headers: dict, test_facility):
        """Amount cells must be numbers (so Excel can sum/sort), not text."""
        r = await client.get("/api/facilities/export.xlsx", headers=auth_headers)
        import openpyxl
        ws = openpyxl.load_workbook(io.BytesIO(r.content))["Facilities"]
        header = [c.value for c in ws[1]]
        amount_col = header.index("amount") + 1
        val = ws.cell(row=2, column=amount_col).value
        assert isinstance(val, (int, float))
