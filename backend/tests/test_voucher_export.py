"""v105 — voucher Excel template export: the spec becomes a real .xlsx laid out
like the printed slip (banner, OUR REF box, signatures, thick frame) with an A4
fit-to-one-page print setup, so filling it later in Excel prints the same."""
import base64
import io

from openpyxl import load_workbook

# a 1x1 red PNG — stands in for the bundled bank logo
_PNG = base64.b64decode(
    "iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAYAAAAfFcSJAAAADUlEQVR4nGP4z8DwHwAFAAH/q842iQAAAABJRU5ErkJggg=="
)
_LOGO = "data:image/png;base64," + base64.b64encode(_PNG).decode()


def _payload():
    return {
        "mode": "irr",
        "logo_png": _LOGO,
        "slips": [
            {"kind": "DEBIT", "title": "SECURITY DOC IRR TD", "date": "07/08/2026",
             "ac_no": "2624-800016-901-090", "amount": "1", "amount_boxed": True,
             "our_ref": "271520 _ STF1260603000001",
             "description": "CHQ NO 536001/237986  FOR IRR 6,383,360,000,000/-   RATE@ IRR 398,960/-",
             "ac_name": "SATIN STAR TRADING LLC",
             "extra_lines": ["LOAN AMOUNT AED 8,000,000/- FOR 48 MONTHS @ 200% LOAN AMOUNT",
                              "IRR CHQ FOR M/s Hani pokht iraninan (karafarin bank - 5300114)"]},
            {"kind": "CREDIT", "title": "SECURITIES", "date": "07/08/2026",
             "ac_no": "2624-869999-901-590", "amount": "1", "amount_boxed": True,
             "our_ref": "271520 _ STF1260603000001",
             "description": "CHQ NO 536001/237986", "ac_name": "SATIN STAR TRADING LLC",
             "extra_lines": []},
        ],
    }


async def test_voucher_excel_export(client, auth_headers):
    r = await client.post("/api/vouchers/export-excel", headers=auth_headers, json=_payload())
    assert r.status_code == 200, r.text
    assert "spreadsheetml" in r.headers["content-type"]
    wb = load_workbook(io.BytesIO(r.content))
    ws = wb["Voucher"]

    # both slips landed with their content
    vals = [str(c.value) for row in ws.iter_rows() for c in row if c.value is not None]
    for needle in ("DEBIT", "CREDIT", "SECURITY DOC IRR TD", "SECURITIES",
                   "2624-800016-901-090", "2624-869999-901-590",
                   "AMOUNT / QUANTITY", "OUR REF :", "Prepared By.", "Authorized Signatures"):
        assert any(needle in v for v in vals), f"missing: {needle}"
    # digit-for-digit IRR line survived
    assert any("6,383,360,000,000" in v for v in vals)
    assert any("karafarin bank - 5300114" in v for v in vals)

    # banner is the lavender merged row
    banner = next(c for row in ws.iter_rows() for c in row if c.value == "SECURITY DOC IRR TD")
    assert banner.fill.fgColor.rgb.endswith("CCCCFF")

    # A4 fit-to-one-page print setup mirroring the web print
    assert str(ws.page_setup.paperSize) == str(ws.PAPERSIZE_A4)
    assert ws.page_setup.orientation == "portrait"
    assert int(ws.page_setup.fitToWidth) == 1 and int(ws.page_setup.fitToHeight) == 1
    assert ws.print_area

    # the logo image is embedded once per slip
    assert len(ws._images) == 2


async def test_voucher_excel_reversal_grid(client, auth_headers):
    """v107 — the reversal mock: centered header stamp, split account row with a
    separate AED cell, and the bordered 2-row info grid instead of OUR REF."""
    payload = {
        "mode": "reversal",
        "logo_png": _LOGO,
        "slips": [{
            "kind": "DEBIT", "title": "PER - CONTRA", "date": "18/08/2026",
            "ac_no": "2624 860185-784-090", "amount": "12,345,678.00",
            "amount_boxed": False, "header_stamp": "SECURITY CHQ REVERSAL",
            "currency_label": "AED",
            "our_ref": "", "description": "", "ac_name": "", "extra_lines": [],
            "grid": [
                [{"t": "Ref. No.", "b": True}, {"t": "BLC1260817000001", "span": 2},
                 {"t": "Borrower Name", "b": True}, {"t": "Maya Ocean"},
                 {"t": "A/c No.", "b": True}, {"t": "115553"}],
                [{"t": "Cheque No.", "b": True}, {"t": "3254856"}, {"t": "by Borrower"},
                 {"t": "Issuer Name", "b": True}, {"t": "ABC Gen. Trd."},
                 {"t": "A/c No.", "b": True}, {"t": "112233"}],
            ],
        }],
    }
    r = await client.post("/api/vouchers/export-excel", headers=auth_headers, json=payload)
    assert r.status_code == 200, r.text
    ws = load_workbook(io.BytesIO(r.content))["Voucher"]
    vals = [str(c.value) for row in ws.iter_rows() for c in row if c.value is not None]
    for needle in ("SECURITY CHQ REVERSAL", "PER - CONTRA", "Account No. :   2624 860185-784-090",
                   "AED", "12,345,678.00", "Ref. No.", "BLC1260817000001", "Borrower Name",
                   "Maya Ocean", "Cheque No.", "3254856", "by Borrower", "Issuer Name",
                   "ABC Gen. Trd.", "112233", "Prepared By."):
        assert any(needle in v for v in vals), f"missing: {needle}"
    assert not any("OUR REF" in v for v in vals)   # the grid replaces the ref box
    # the grid's label cells are bold and bordered
    ref = next(c for row in ws.iter_rows() for c in row if c.value == "Ref. No.")
    assert ref.font.bold and ref.border.top.style == "thin"


async def test_voucher_excel_requires_auth(client):
    r = await client.post("/api/vouchers/export-excel", json=_payload())
    assert r.status_code in (401, 403)


async def test_voucher_excel_rejects_bad_logo(client, auth_headers):
    p = _payload()
    p["logo_png"] = "data:text/html;base64,PGh0bWw+"   # not an image → silently no logo
    r = await client.post("/api/vouchers/export-excel", headers=auth_headers, json=p)
    assert r.status_code == 200
    ws = load_workbook(io.BytesIO(r.content))["Voucher"]
    assert len(ws._images) == 0
