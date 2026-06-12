#!/usr/bin/env python3
"""Convert the core-banking "Customer Listing" export into the merge dataset.

The bank periodically exports its full customer base as an Excel workbook
(``CustomerListing_YYYYMMDD_HHMMSS.xlsx``, ~60k rows, one sheet ``Report`` with
a header on row 6). This script distils that export into the compact, gzipped
JSON the panel merges at startup
(:mod:`app.services.data_merge` → ``_merge_customer_listing_*``), writing
``backend/app/data/merge/customer_listing.jsonl.gz`` (newline-delimited JSON,
gzipped, so the merge can stream it record-by-record within the 512MB instance).

Business rules (confirmed with the data owner):

* **Account number** = column ``CUSTOMER NUMBER``. A real account number is
  *exactly 6 digits*. Anything shorter/longer (blank, 4-digit branch echoes such
  as ``2533``, 15/24-digit IBAN-like strings, ``1`` …) is **not** an account and
  is dropped.
* **Branch** = column ``BRANCH CODE`` (4 digits). It is mapped to a human label
  and stored as ``"<Name> (<code>)"`` for the panel; the raw code is kept too.
  Codes outside the known set (only ``3535`` in this export — head-office /
  correspondent-bank accounts) are labelled ``"Head Office (3535)"``.
* **Account type** is derived from ``ENTITY TYPE DESC``.

The raw workbook is intentionally *not* committed (large + contains KYC PII).
Drop it somewhere local and pass its path:

    python scripts/generate_customer_listing.py /path/to/CustomerListing_*.xlsx

Re-running is safe and deterministic; the merge step is idempotent regardless.
"""
from __future__ import annotations

import datetime
import gzip
import json
import re
import sys
from collections import Counter
from pathlib import Path

import openpyxl

# 4-digit branch code -> readable name. The panel shows "<Name> (<code>)".
BRANCH_NAMES = {
    "2533": "Bur Dubai",
    "2690": "Abu Dhabi",
    "2776": "Sharjah",
    "2900": "Ajman",
    "4350": "Sheikh Zayed Road",
    "2624": "Al Maktoum",  # the main / central branch
    "2898": "Murshid Bazar",
    "1741": "Al Ain",
    "3535": "Head Office",  # head-office / correspondent-bank accounts
}

# ENTITY TYPE DESC -> Customer.account_type enum value (retail|corporate|sme).
ENTITY_TO_TYPE = {
    "individual": "retail",
    "joint account": "retail",
    "company": "corporate",
    "partnership": "sme",
}

_SIX_DIGITS = re.compile(r"^\d{6}$")

# 0-based column indices on the ``Report`` sheet (header row 6, data from row 7).
COL = dict(
    BRM=0,
    ENTITY_TYPE=1,
    ENTITY_TYPE_DESC=2,
    FULL_NAME=3,
    ARABIC_NAME=4,
    FIRST=5,
    MIDDLE=6,
    LAST=7,
    COMPANY=8,
    CUSTOMER_TYPE=9,
    CUSTOMER_TYPE_DESC=10,
    NIC=11,
    PASSPORT=12,
    TRADE_LICENSE=13,
    BRANCH_CODE=14,
    STATUS=15,
    STATUS_DESC=16,
    DATE_ADDED=17,
    LMNT_DATE=18,
    MOBILE=19,
    EMAIL=20,
    NATIONALITY=21,
    CUSTOMER_NUMBER=22,
    RR_PEP=23,
    PEP_DESC=24,
    SIMILARITY=25,
)

# Excel epoch placeholders that mean "no real date".
_EPOCH_PLACEHOLDERS = {"1900-01-01", "1899-12-31", "1899-12-30"}

OUT = (
    Path(__file__).resolve().parent.parent
    / "backend"
    / "app"
    / "data"
    / "merge"
    / "customer_listing.jsonl.gz"
)


def _s(v) -> str:
    """Normalise a cell value to a trimmed string (ints without a trailing .0)."""
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    if isinstance(v, datetime.datetime):
        return v.date().isoformat()
    return str(v).strip()


def build_records(xlsx_path: Path) -> tuple[list[dict], Counter, dict]:
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]

    records: list[dict] = []
    seen: set[str] = set()
    stats = dict(total=0, kept=0, dropped_not_6=0, dropped_dup=0)
    branches: Counter = Counter()

    for row in ws.iter_rows(min_row=7, values_only=True):
        if row is None or all(c is None for c in row):
            continue
        stats["total"] += 1

        acc = _s(row[COL["CUSTOMER_NUMBER"]])
        if not _SIX_DIGITS.match(acc):
            stats["dropped_not_6"] += 1
            continue
        if acc in seen:
            stats["dropped_dup"] += 1
            continue
        seen.add(acc)

        full = _s(row[COL["FULL_NAME"]])
        company = _s(row[COL["COMPANY"]])
        parts = [
            p
            for p in (
                _s(row[COL["FIRST"]]),
                _s(row[COL["MIDDLE"]]),
                _s(row[COL["LAST"]]),
            )
            if p and p != "."
        ]
        name = full or company or " ".join(parts) or f"Account {acc}"

        code = _s(row[COL["BRANCH_CODE"]])
        bname = BRANCH_NAMES.get(code)
        label = f"{bname} ({code})" if bname else code

        entity = _s(row[COL["ENTITY_TYPE_DESC"]])
        atype = ENTITY_TO_TYPE.get(entity.lower(), "retail")
        branches[code] += 1

        added = _s(row[COL["DATE_ADDED"]])
        if added in _EPOCH_PLACEHOLDERS:
            added = ""

        rec = {"account_no": acc, "name": name[:200], "account_type": atype}

        def put(key, value, limit=None):
            value = (value or "").strip()
            if value:
                rec[key] = value[:limit] if limit else value

        put("name_ar", _s(row[COL["ARABIC_NAME"]]), 200)
        put("branch_label", label, 100)
        put("branch_code", code, 20)
        put("entity_type", entity, 40)
        put("customer_type", _s(row[COL["CUSTOMER_TYPE_DESC"]]), 60)
        put("email", _s(row[COL["EMAIL"]]), 100)
        put("mobile", _s(row[COL["MOBILE"]]), 50)
        put("nationality", _s(row[COL["NATIONALITY"]]), 80)
        put("passport_no", _s(row[COL["PASSPORT"]]), 80)
        put("national_id", _s(row[COL["NIC"]]), 80)
        put("trade_license_no", _s(row[COL["TRADE_LICENSE"]]), 80)
        put("status_desc", _s(row[COL["STATUS_DESC"]]), 50)
        put("pep_status", _s(row[COL["PEP_DESC"]]), 40)
        put("rr_pep", _s(row[COL["RR_PEP"]]), 10)
        put("brm_code", _s(row[COL["BRM"]]), 30)
        put("date_added", added, 30)

        records.append(rec)

    wb.close()
    stats["kept"] = len(records)
    return records, branches, stats


def main() -> int:
    if len(sys.argv) < 2:
        print(
            "usage: generate_customer_listing.py <CustomerListing.xlsx>",
            file=sys.stderr,
        )
        return 2
    xlsx_path = Path(sys.argv[1])
    if not xlsx_path.exists():
        print(f"file not found: {xlsx_path}", file=sys.stderr)
        return 2

    records, branches, stats = build_records(xlsx_path)
    # Write newline-delimited JSON (one record per line), gzipped. This lets the
    # merge step stream the file record-by-record instead of loading all ~44k
    # dicts at once — essential to stay within the 512MB production instance.
    OUT.parent.mkdir(parents=True, exist_ok=True)
    raw_len = 0
    with gzip.open(OUT, "wt", encoding="utf-8", compresslevel=9) as fh:
        for rec in records:
            line = json.dumps(rec, ensure_ascii=False, separators=(",", ":"))
            raw_len += len(line.encode("utf-8")) + 1
            fh.write(line + "\n")

    print(f"source rows         : {stats['total']}")
    print(f"kept (6-digit accts): {stats['kept']}")
    print(f"dropped (not 6-digit): {stats['dropped_not_6']}")
    print(f"dropped (duplicate)  : {stats['dropped_dup']}")
    print(f"branches             : {dict(branches.most_common())}")
    gz_mb = OUT.stat().st_size / 1024 / 1024
    raw_mb = raw_len / 1024 / 1024
    print(f"wrote {OUT}")
    print(f"  {gz_mb:.2f} MB gz, {raw_mb:.1f} MB raw, {len(records)} lines")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
